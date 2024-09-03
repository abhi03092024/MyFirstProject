import openpyxl

#file--->workbook--->Sheets---->rows---->cells

file = "C:\Data.xlsx"
workbook = openpyxl.load_workbook(file)
Sheet1 = workbook["Sheet1"]

rows = Sheet1.max_row # For count all the rows from excel sheet
cols = Sheet1.max_column # For count all the column from the excel sheet

#How to read dat from excel
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(Sheet1.cell(r, c).value,end="       ")
    print()